"""
Анализ структуры шаблона для выгрузки прогноза
"""
import openpyxl
from pathlib import Path

def analyze_forecast_template():
    """Анализирует структуру Excel шаблона для прогноза"""

    template_path = Path(__file__).parent.parent.parent / "xls" / "Планирование_10.2025-3.xlsx"

    if not template_path.exists():
        print(f"Файл не найден: {template_path}")
        return

    print(f"Анализ файла: {template_path}\n")

    # Открываем файл
    wb = openpyxl.load_workbook(template_path, data_only=False)

    print(f"Всего листов: {len(wb.sheetnames)}")
    print(f"Листы: {wb.sheetnames}\n")

    # Анализируем каждый лист
    for sheet_name in wb.sheetnames:
        print(f"\n{'='*80}")
        print(f"ЛИСТ: {sheet_name}")
        print(f"{'='*80}")

        ws = wb[sheet_name]

        # Размеры листа
        print(f"Максимальная строка: {ws.max_row}")
        print(f"Максимальный столбец: {ws.max_column}")

        # Выводим первые 20 строк и первые 10 колонок
        print("\nПервые 20 строк (A-J):")
        print("-" * 120)

        for row_idx in range(1, min(21, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(11, ws.max_column + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = cell.value

                # Форматируем значение для вывода
                if value is None:
                    display_value = ""
                elif isinstance(value, str):
                    display_value = value[:20]  # Обрезаем длинные строки
                else:
                    display_value = str(value)[:20]

                row_data.append(display_value)

            print(f"Row {row_idx:2d}: {' | '.join(f'{v:20s}' for v in row_data)}")

        # Ищем merged cells
        if ws.merged_cells:
            print(f"\nОбъединенные ячейки:")
            for merged_range in list(ws.merged_cells.ranges)[:10]:  # Первые 10
                print(f"  {merged_range}")

        # Проверяем наличие формул
        formula_cells = []
        for row in ws.iter_rows(min_row=1, max_row=min(20, ws.max_row)):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula_cells.append((cell.coordinate, cell.value))

        if formula_cells:
            print(f"\nФормулы (первые 10):")
            for coord, formula in formula_cells[:10]:
                print(f"  {coord}: {formula[:50]}")

    wb.close()
    print("\n" + "="*80)
    print("Анализ завершен")

if __name__ == "__main__":
    analyze_forecast_template()
