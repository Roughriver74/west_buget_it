"""
Скрипт для импорта данных ФОТ (Фонд оплаты труда) из Excel файла
"""

import sys
import os
from datetime import datetime, date
from decimal import Decimal

# Добавляем путь к приложению
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from sqlalchemy.orm import Session
from openpyxl import load_workbook

from app.db.session import SessionLocal
from app.db.models import Employee, Payroll, Organization, EmployeeStatusEnum, PositionLevelEnum


def parse_date(date_str):
    """Парсинг даты из разных форматов"""
    if isinstance(date_str, date):
        return date_str
    if isinstance(date_str, datetime):
        return date_str.date()
    if isinstance(date_str, str):
        # Пробуем разные форматы
        for fmt in ['%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y']:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue
    return None


def import_employees_from_excel(file_path: str, sheet_name: str = None, db: Session = None):
    """
    Импорт сотрудников из Excel файла

    Ожидаемая структура листа:
    - Колонка A: ФИО
    - Колонка B: Должность
    - Колонка C: Оклад (₽/мес)
    - Колонка D: Дата приема
    - Колонка E: Статус (Работает/Уволен)
    - Колонка F: Email (опционально)
    - Колонка G: Телефон (опционально)
    - Колонка H: Организация (опционально)
    """

    close_db = False
    if db is None:
        db = SessionLocal()
        close_db = True

    try:
        # Загружаем Excel файл
        wb = load_workbook(file_path, data_only=True)

        # Выбираем лист
        if sheet_name:
            ws = wb[sheet_name]
        else:
            # Если не указан лист, берем 3-й (индекс 2)
            if len(wb.sheetnames) >= 3:
                ws = wb[wb.sheetnames[2]]
            else:
                ws = wb.active

        print(f"Читаем лист: {ws.title}")
        print(f"Максимальная строка: {ws.max_row}")

        # Получаем организацию по умолчанию
        default_org = db.query(Organization).first()

        # Статистика
        created_count = 0
        updated_count = 0
        skipped_count = 0

        # Начинаем с 2-й строки (пропускаем заголовок)
        for row_idx in range(2, ws.max_row + 1):
            row = ws[row_idx]

            # Читаем значения
            full_name = row[0].value  # A: ФИО
            if not full_name or str(full_name).strip() == "":
                skipped_count += 1
                continue

            position = row[1].value if len(row) > 1 and row[1].value else "IT-специалист"  # B: Должность
            base_salary = row[2].value if len(row) > 2 and row[2].value else 100000  # C: Оклад
            hire_date_str = row[3].value if len(row) > 3 else None  # D: Дата приема
            status_str = row[4].value if len(row) > 4 else "Работает"  # E: Статус
            email = row[5].value if len(row) > 5 else None  # F: Email
            phone = row[6].value if len(row) > 6 else None  # G: Телефон
            org_name = row[7].value if len(row) > 7 else None  # H: Организация

            # Парсим дату приема
            hire_date = parse_date(hire_date_str)
            if not hire_date:
                hire_date = date.today()

            # Определяем статус
            status = EmployeeStatusEnum.ACTIVE
            if status_str and isinstance(status_str, str):
                if "уволен" in status_str.lower() or "dismissed" in status_str.lower():
                    status = EmployeeStatusEnum.DISMISSED
                elif "отпуск" in status_str.lower() or "vacation" in status_str.lower():
                    status = EmployeeStatusEnum.VACATION

            # Ищем организацию
            organization = None
            if org_name:
                organization = db.query(Organization).filter(
                    Organization.name.ilike(f"%{org_name}%")
                ).first()
            if not organization:
                organization = default_org

            # Преобразуем зарплату
            try:
                salary = Decimal(str(base_salary))
            except:
                salary = Decimal('100000')

            # Проверяем, существует ли сотрудник
            existing = db.query(Employee).filter(Employee.full_name == str(full_name).strip()).first()

            if existing:
                # Обновляем существующего
                existing.position = str(position).strip() if position else existing.position
                existing.base_salary = salary
                existing.hire_date = hire_date
                existing.status = status
                existing.email = str(email).strip() if email else existing.email
                existing.phone = str(phone).strip() if phone else existing.phone
                existing.organization_id = organization.id if organization else existing.organization_id

                updated_count += 1
                print(f"  Обновлен: {full_name} - {position} - {salary:,.0f} ₽")
            else:
                # Создаем нового сотрудника
                employee = Employee(
                    full_name=str(full_name).strip(),
                    position=str(position).strip() if position else "IT-специалист",
                    base_salary=salary,
                    hire_date=hire_date,
                    status=status,
                    email=str(email).strip() if email else None,
                    phone=str(phone).strip() if phone else None,
                    organization_id=organization.id if organization else None,
                    tax_rate=Decimal('30.0')  # Стандартная ставка налогов
                )
                db.add(employee)
                created_count += 1
                print(f"  Создан: {full_name} - {position} - {salary:,.0f} ₽")

        # Сохраняем изменения
        db.commit()

        print("\n" + "="*60)
        print(f"✅ Импорт завершен!")
        print(f"  Создано: {created_count}")
        print(f"  Обновлено: {updated_count}")
        print(f"  Пропущено: {skipped_count}")
        print("="*60)

        return {
            "created": created_count,
            "updated": updated_count,
            "skipped": skipped_count
        }

    except Exception as e:
        print(f"❌ Ошибка при импорте: {e}")
        import traceback
        traceback.print_exc()
        db.rollback()
        return None
    finally:
        if close_db:
            db.close()


def generate_payrolls_for_year(year: int, db: Session = None):
    """
    Генерация записей зарплат на весь год для всех активных сотрудников
    """
    close_db = False
    if db is None:
        db = SessionLocal()
        close_db = True

    try:
        # Получаем всех активных сотрудников
        employees = db.query(Employee).filter(
            Employee.status == EmployeeStatusEnum.ACTIVE
        ).all()

        print(f"\n📊 Генерация записей зарплат на {year} год для {len(employees)} сотрудников")

        created_count = 0

        for employee in employees:
            for month in range(1, 13):  # Январь - Декабрь
                # Проверяем, есть ли уже запись
                existing = db.query(Payroll).filter(
                    Payroll.employee_id == employee.id,
                    Payroll.year == year,
                    Payroll.month == month
                ).first()

                if existing:
                    continue

                # Расчеты
                base_salary = employee.base_salary
                bonus = Decimal('0')  # Можно добавить логику премий
                other_payments = Decimal('0')

                gross_salary = base_salary + bonus + other_payments

                # Налоги с сотрудника (НДФЛ 13%)
                taxes = gross_salary * Decimal('0.13')
                net_salary = gross_salary - taxes

                # Налоги работодателя (ЕСН ~30%)
                employer_taxes = gross_salary * (employee.tax_rate / 100)
                total_cost = gross_salary + employer_taxes

                # Создаем запись
                payroll = Payroll(
                    employee_id=employee.id,
                    year=year,
                    month=month,
                    base_salary=base_salary,
                    bonus=bonus,
                    other_payments=other_payments,
                    gross_salary=gross_salary,
                    taxes=taxes,
                    net_salary=net_salary,
                    employer_taxes=employer_taxes,
                    total_cost=total_cost,
                    worked_days=None  # Заполняется позже
                )
                db.add(payroll)
                created_count += 1

        db.commit()

        print(f"✅ Создано {created_count} записей зарплат")

        return created_count

    except Exception as e:
        print(f"❌ Ошибка при генерации зарплат: {e}")
        db.rollback()
        return 0
    finally:
        if close_db:
            db.close()


if __name__ == "__main__":
    print("="*60)
    print("ИМПОРТ ДАННЫХ ФОТ ИЗ EXCEL")
    print("="*60)

    # Путь к файлу
    file_path = "../xls/Планфакт2025.xlsx"

    # Импортируем сотрудников
    result = import_employees_from_excel(file_path, sheet_name=None)  # sheet_name=None использует 3-й лист

    if result and (result['created'] > 0 or result['updated'] > 0):
        # Генерируем зарплаты на 2025 год
        print("\n")
        generate_payrolls_for_year(2025)

    print("\n✅ Готово!")
