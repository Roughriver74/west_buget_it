"""
Анализ структуры файла ДДС (движение денежных средств)
"""

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl

# Полный путь к файлу
file_path = '/Users/evgenijsikunov/projects/acme/acme_buget_it/xls/Копия ДДС_2025_без доп вкладок_03.10.2025.xlsx'

print('=' * 80)
print('СТРУКТУРА ФАЙЛА ДДС 2025')
print('=' * 80)

wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

print(f'\n📊 Количество листов: {len(wb.sheetnames)}')
print(f'Листы: {wb.sheetnames}\n')

for sheet_name in wb.sheetnames[:5]:  # Первые 5 листов
    print('\n' + '=' * 80)
    print(f'📄 Лист: {sheet_name}')
    print('=' * 80)

    ws = wb[sheet_name]

    # Размеры
    print(f'Размер: {ws.max_row} строк x {ws.max_column} колонок')

    # Первые 15 строк
    print(f'\nПервые 15 строк:')
    print('-' * 80)

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), 1):
        # Показать только первые 20 колонок
        row_data = []
        for cell in row[:20]:
            if cell is not None:
                cell_str = str(cell)
                if len(cell_str) > 40:
                    cell_str = cell_str[:37] + '...'
                row_data.append(cell_str)
            else:
                row_data.append('')

        # Показать только если есть данные
        if any(row_data):
            print(f'{row_idx:3d}: {" | ".join(row_data)}')

    # Найти заголовки (строка с максимальным количеством непустых ячеек)
    print(f'\n🔍 Поиск строки с заголовками...')
    max_filled = 0
    header_row = 1

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
        filled = sum(1 for cell in row[:30] if cell is not None and str(cell).strip())
        if filled > max_filled:
            max_filled = filled
            header_row = row_idx

    print(f'Вероятная строка заголовков: {header_row} ({max_filled} заполненных ячеек)')

    # Показать заголовки
    headers = []
    for cell in ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True).__next__():
        if cell is not None:
            headers.append(str(cell))
        else:
            headers.append('')

    print('\nЗаголовки:')
    for idx, header in enumerate(headers[:25], 1):
        if header.strip():
            print(f'  {idx:2d}. {header}')

wb.close()

print('\n' + '=' * 80)
print('ЗАВЕРШЕНО')
print('=' * 80)
