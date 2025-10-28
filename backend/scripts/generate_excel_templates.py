"""
Script to generate Excel templates for import functionality
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

# Create templates directory if it doesn't exist
TEMPLATES_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'templates')
os.makedirs(TEMPLATES_DIR, exist_ok=True)

# Styles
header_font = Font(bold=True, color="FFFFFF", size=12)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")
example_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def auto_adjust_column_width(worksheet):
    """Auto-adjust column widths based on content"""
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def apply_header_style(worksheet, row=1):
    """Apply header style to the first row"""
    for cell in worksheet[row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border


def create_categories_template():
    """Шаблон для импорта категорий бюджета"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Категории"

    # Headers
    headers = ["Название", "Тип", "Описание", "Родитель"]
    ws.append(headers)
    apply_header_style(ws)

    # Example data
    examples = [
        ["Программное обеспечение", "OPEX", "Закупка и лицензирование ПО", ""],
        ["Серверы", "CAPEX", "Покупка серверного оборудования", ""],
        ["Microsoft Office 365", "OPEX", "Подписка на офисный пакет", "Программное обеспечение"],
        ["Хостинг и облако", "OPEX", "Аренда облачных сервисов", ""],
    ]

    for row in examples:
        ws.append(row)

    # Apply example style
    for row in range(2, len(examples) + 2):
        for cell in ws[row]:
            cell.fill = example_fill
            cell.border = thin_border

    auto_adjust_column_width(ws)

    # Add notes
    ws.append([])
    ws.append(["ИНСТРУКЦИЯ:"])
    ws.append(["1. Название - обязательное поле, название категории расходов"])
    ws.append(["2. Тип - обязательное поле, OPEX или CAPEX"])
    ws.append(["3. Описание - опциональное, описание категории"])
    ws.append(["4. Родитель - опциональное, название родительской категории для вложенности"])
    ws.append(["5. Серые строки - примеры данных, их можно удалить"])

    file_path = os.path.join(TEMPLATES_DIR, "template_categories.xlsx")
    wb.save(file_path)
    print(f"✓ Created: {file_path}")


def create_contractors_template():
    """Шаблон для импорта контрагентов"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Контрагенты"

    # Headers
    headers = ["Название", "ИНН", "Контакты"]
    ws.append(headers)
    apply_header_style(ws)

    # Example data
    examples = [
        ["ООО Рога и Копыта", "7701234567", "info@rogaikopyta.ru, +7(495)123-45-67"],
        ["АО СофтСистемс", "7707654321", "sales@softsystems.ru"],
        ["ИП Иванов И.И.", "123456789012", "+7(903)555-66-77"],
    ]

    for row in examples:
        ws.append(row)

    for row in range(2, len(examples) + 2):
        for cell in ws[row]:
            cell.fill = example_fill
            cell.border = thin_border

    auto_adjust_column_width(ws)

    ws.append([])
    ws.append(["ИНСТРУКЦИЯ:"])
    ws.append(["1. Название - обязательное поле"])
    ws.append(["2. ИНН - опциональное, используется для поиска дублей"])
    ws.append(["3. Контакты - опциональное, email, телефон"])
    ws.append(["4. Дубли определяются по ИНН (если указан) или по названию"])

    file_path = os.path.join(TEMPLATES_DIR, "template_contractors.xlsx")
    wb.save(file_path)
    print(f"✓ Created: {file_path}")


def create_organizations_template():
    """Шаблон для импорта организаций"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Организации"

    # Headers
    headers = ["Название"]
    ws.append(headers)
    apply_header_style(ws)

    # Example data
    examples = [
        ["ООО Наша Компания"],
        ["ООО Дочерняя Компания 1"],
        ["ООО Дочерняя Компания 2"],
    ]

    for row in examples:
        ws.append(row)

    for row in range(2, len(examples) + 2):
        for cell in ws[row]:
            cell.fill = example_fill
            cell.border = thin_border

    auto_adjust_column_width(ws)

    ws.append([])
    ws.append(["ИНСТРУКЦИЯ:"])
    ws.append(["1. Название - обязательное поле"])
    ws.append(["2. Это название вашей организации (плательщика)"])

    file_path = os.path.join(TEMPLATES_DIR, "template_organizations.xlsx")
    wb.save(file_path)
    print(f"✓ Created: {file_path}")


def create_payroll_plans_template():
    """Шаблон для импорта планов зарплаты"""
    wb = Workbook()
    ws = wb.active
    ws.title = "План ФОТ"

    # Headers
    headers = ["Год", "Месяц", "Сотрудник", "Оклад", "Премия месячная", "Премия квартальная", "Премия годовая", "Прочие выплаты", "Примечания"]
    ws.append(headers)
    apply_header_style(ws)

    # Example data
    current_year = datetime.now().year
    examples = [
        [current_year, 1, "Иванов Иван Иванович", 150000, 30000, 0, 0, 0, ""],
        [current_year, 1, "Петров Петр Петрович", 120000, 24000, 0, 0, 0, ""],
        [current_year, 2, "Иванов Иван Иванович", 150000, 30000, 0, 0, 0, ""],
        [current_year, 3, "Иванов Иван Иванович", 150000, 30000, 50000, 0, 0, "Квартальная премия"],
    ]

    for row in examples:
        ws.append(row)

    for row in range(2, len(examples) + 2):
        for cell in ws[row]:
            cell.fill = example_fill
            cell.border = thin_border

    auto_adjust_column_width(ws)

    ws.append([])
    ws.append(["ИНСТРУКЦИЯ:"])
    ws.append(["1. Год - обязательное, год плана (число)"])
    ws.append(["2. Месяц - обязательное, номер месяца 1-12"])
    ws.append(["3. Сотрудник - обязательное, ФИО сотрудника"])
    ws.append(["4. Оклад - обязательное, базовый оклад"])
    ws.append(["5. Премия месячная - опциональное, месячная премия"])
    ws.append(["6. Премия квартальная - опциональное, квартальная премия"])
    ws.append(["7. Премия годовая - опциональное, годовая премия"])
    ws.append(["8. Прочие выплаты - опциональное"])
    ws.append(["9. Примечания - опциональное"])
    ws.append(["10. Сотрудник будет создан автоматически, если не существует"])

    file_path = os.path.join(TEMPLATES_DIR, "template_payroll_plans.xlsx")
    wb.save(file_path)
    print(f"✓ Created: {file_path}")


def create_kpi_template():
    """Шаблон для импорта КПИ"""
    wb = Workbook()
    ws = wb.active
    ws.title = "УПРАВЛЕНИЕ КПИ"

    # Title
    ws['A1'] = "СИСТЕМА УПРАВЛЕНИЯ КПИ ИТ ОТДЕЛА"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:G1')

    ws['A3'] = f"Период: {datetime.now().strftime('%B %Y')}"
    ws['A3'].font = Font(italic=True)

    # Headers at row 6
    ws['A5'] = "ОСНОВНАЯ ИНФОРМАЦИЯ"
    ws['A5'].font = Font(bold=True, size=12)

    headers = ["Сотрудник", "Оклад", "Должность", "ЗП Год", "Базовая премия", "Вариант премии", "КПИ Общий %"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Example data
    examples = [
        ["Иванов Иван Иванович", 200000, "Руководитель ИТ", "=B7*12", 80000, "Результативный", 100.0],
        ["Петров Петр Петрович", 150000, "Разработчик", "=B8*12", 50000, "Результативный", 95.0],
        ["Сидоров Сидор Сидорович", 130000, "Системный администратор", "=B9*12", 30000, "Фиксированный", 100.0],
        ["Васильева Анна Петровна", 120000, "Аналитик", "=B10*12", 40000, "Смешанный", 105.0],
    ]

    for idx, row in enumerate(examples, start=7):
        for col, value in enumerate(row, start=1):
            cell = ws.cell(row=idx, column=col)
            cell.value = value
            cell.fill = example_fill
            cell.border = thin_border

    auto_adjust_column_width(ws)

    # Instructions
    ws['A13'] = "ИНСТРУКЦИЯ:"
    ws['A13'].font = Font(bold=True)
    instructions = [
        "1. НЕ ИЗМЕНЯЙТЕ название листа - должно быть 'УПРАВЛЕНИЕ КПИ'",
        "2. НЕ ИЗМЕНЯЙТЕ расположение таблицы - заголовки должны быть в строке 6",
        "3. Столбец A (Сотрудник) - обязательно, ФИО",
        "4. Столбец B (Оклад) - обязательно, базовая зарплата",
        "5. Столбец C (Должность) - обязательно, должность сотрудника",
        "6. Столбец D (ЗП Год) - можно оставить формулу =B*12",
        "7. Столбец E (Базовая премия) - обязательно, база для расчета премии",
        "8. Столбец F (Вариант премии) - Результативный/Фиксированный/Смешанный",
        "9. Столбец G (КПИ Общий %) - процент выполнения КПИ (0-200%)",
        "10. Сотрудники будут созданы или обновлены автоматически",
    ]

    for idx, instruction in enumerate(instructions, start=14):
        ws[f'A{idx}'] = instruction

    file_path = os.path.join(TEMPLATES_DIR, "template_kpi.xlsx")
    wb.save(file_path)
    print(f"✓ Created: {file_path}")


def main():
    """Generate all Excel templates"""
    print("Generating Excel templates...")
    print(f"Output directory: {TEMPLATES_DIR}\n")

    create_categories_template()
    create_contractors_template()
    create_organizations_template()
    create_payroll_plans_template()
    create_kpi_template()

    print(f"\n✅ All templates generated successfully in {TEMPLATES_DIR}")


if __name__ == "__main__":
    main()
