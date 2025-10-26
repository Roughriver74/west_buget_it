"""
Excel export utilities for IT Budget Manager
"""

from io import BytesIO
from typing import List, Dict, Any, Optional
from datetime import datetime, date
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from calendar import monthrange


class ExcelExporter:
    """Helper class for exporting data to Excel"""

    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

    OPEX_FILL = PatternFill(start_color="E7E6FF", end_color="E7E6FF", fill_type="solid")
    CAPEX_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    TOTAL_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    TOTAL_FONT = Font(bold=True, size=11)

    BORDER = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    @staticmethod
    def export_budget_plan(year: int, categories_data: List[Dict[str, Any]]) -> BytesIO:
        """
        Export budget plan to Excel

        Args:
            year: Year of the budget plan
            categories_data: List of category data with monthly breakdown

        Returns:
            BytesIO: Excel file in memory
        """
        wb = Workbook()
        ws = wb.active
        ws.title = f"План {year}"

        # Header
        ws['A1'] = f"Бюджетный план на {year} год"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:N1')

        # Column headers
        headers = ['Статья расходов', 'Тип'] + \
                  ['Янв', 'Фев', 'Мар', 'Апр', 'Май', 'Июн', 'Июл', 'Авг', 'Сен', 'Окт', 'Ноя', 'Дек'] + \
                  ['ИТОГО']

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num, value=header)
            cell.font = ExcelExporter.HEADER_FONT
            cell.fill = ExcelExporter.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = ExcelExporter.BORDER

        # Data rows
        current_row = 4
        opex_total_row_start = current_row
        capex_total_row_start = None

        # OPEX categories
        for cat in [c for c in categories_data if c['category_type'] == 'OPEX']:
            ws.cell(row=current_row, column=1, value=cat['category_name'])
            ws.cell(row=current_row, column=2, value='OPEX')

            # Apply OPEX styling
            for col in range(1, 15):
                ws.cell(row=current_row, column=col).fill = ExcelExporter.OPEX_FILL
                ws.cell(row=current_row, column=col).border = ExcelExporter.BORDER

            # Monthly values
            row_total = 0
            for month in range(1, 13):
                month_str = str(month)
                value = cat['months'].get(month_str, {}).get('planned_amount', 0)
                ws.cell(row=current_row, column=month + 2, value=value)
                ws.cell(row=current_row, column=month + 2).number_format = '#,##0.00'
                row_total += value

            # Row total
            ws.cell(row=current_row, column=14, value=row_total)
            ws.cell(row=current_row, column=14).number_format = '#,##0.00'
            ws.cell(row=current_row, column=14).font = Font(bold=True)

            current_row += 1

        opex_total_row_end = current_row - 1

        # OPEX Total row
        ws.cell(row=current_row, column=1, value='ИТОГО OPEX')
        ws.cell(row=current_row, column=1).font = ExcelExporter.TOTAL_FONT
        ws.cell(row=current_row, column=2, value='OPEX')

        for col in range(1, 15):
            ws.cell(row=current_row, column=col).fill = ExcelExporter.TOTAL_FILL
            ws.cell(row=current_row, column=col).border = ExcelExporter.BORDER
            ws.cell(row=current_row, column=col).font = ExcelExporter.TOTAL_FONT

        # Calculate OPEX totals
        for month in range(1, 13):
            col_letter = get_column_letter(month + 2)
            formula = f"=SUM({col_letter}{opex_total_row_start}:{col_letter}{opex_total_row_end})"
            ws.cell(row=current_row, column=month + 2, value=formula)
            ws.cell(row=current_row, column=month + 2).number_format = '#,##0.00'

        # OPEX grand total
        ws.cell(row=current_row, column=14, value=f"=SUM(C{current_row}:N{current_row})")
        ws.cell(row=current_row, column=14).number_format = '#,##0.00'

        current_row += 2
        capex_total_row_start = current_row

        # CAPEX categories
        for cat in [c for c in categories_data if c['category_type'] == 'CAPEX']:
            ws.cell(row=current_row, column=1, value=cat['category_name'])
            ws.cell(row=current_row, column=2, value='CAPEX')

            # Apply CAPEX styling
            for col in range(1, 15):
                ws.cell(row=current_row, column=col).fill = ExcelExporter.CAPEX_FILL
                ws.cell(row=current_row, column=col).border = ExcelExporter.BORDER

            # Monthly values
            row_total = 0
            for month in range(1, 13):
                month_str = str(month)
                value = cat['months'].get(month_str, {}).get('planned_amount', 0)
                ws.cell(row=current_row, column=month + 2, value=value)
                ws.cell(row=current_row, column=month + 2).number_format = '#,##0.00'
                row_total += value

            # Row total
            ws.cell(row=current_row, column=14, value=row_total)
            ws.cell(row=current_row, column=14).number_format = '#,##0.00'
            ws.cell(row=current_row, column=14).font = Font(bold=True)

            current_row += 1

        capex_total_row_end = current_row - 1

        # CAPEX Total row
        if capex_total_row_start <= capex_total_row_end:
            ws.cell(row=current_row, column=1, value='ИТОГО CAPEX')
            ws.cell(row=current_row, column=1).font = ExcelExporter.TOTAL_FONT
            ws.cell(row=current_row, column=2, value='CAPEX')

            for col in range(1, 15):
                ws.cell(row=current_row, column=col).fill = ExcelExporter.TOTAL_FILL
                ws.cell(row=current_row, column=col).border = ExcelExporter.BORDER
                ws.cell(row=current_row, column=col).font = ExcelExporter.TOTAL_FONT

            # Calculate CAPEX totals
            for month in range(1, 13):
                col_letter = get_column_letter(month + 2)
                formula = f"=SUM({col_letter}{capex_total_row_start}:{col_letter}{capex_total_row_end})"
                ws.cell(row=current_row, column=month + 2, value=formula)
                ws.cell(row=current_row, column=month + 2).number_format = '#,##0.00'

            # CAPEX grand total
            ws.cell(row=current_row, column=14, value=f"=SUM(C{current_row}:N{current_row})")
            ws.cell(row=current_row, column=14).number_format = '#,##0.00'

        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 10
        for col in range(3, 15):
            ws.column_dimensions[get_column_letter(col)].width = 12

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    @staticmethod
    def export_expenses(expenses: List[Dict[str, Any]], filters: Dict[str, Any] = None) -> BytesIO:
        """
        Export expenses to Excel

        Args:
            expenses: List of expense records
            filters: Applied filters (for report header)

        Returns:
            BytesIO: Excel file in memory
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Расходы"

        # Header
        title = "Отчет по расходам"
        if filters:
            filter_parts = []
            if filters.get('year'):
                filter_parts.append(f"Год: {filters['year']}")
            if filters.get('month'):
                filter_parts.append(f"Месяц: {filters['month']}")
            if filters.get('category'):
                filter_parts.append(f"Категория: {filters['category']}")
            if filter_parts:
                title += f" ({', '.join(filter_parts)})"

        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:J1')

        # Column headers
        headers = [
            'Номер',
            'Дата заявки',
            'Категория',
            'Тип',
            'Контрагент',
            'Организация',
            'Сумма',
            'Статус',
            'Дата оплаты',
            'Комментарий'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num, value=header)
            cell.font = ExcelExporter.HEADER_FONT
            cell.fill = ExcelExporter.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = ExcelExporter.BORDER

        # Data rows
        current_row = 4
        for expense in expenses:
            ws.cell(row=current_row, column=1, value=expense.get('number', ''))

            # Format request date
            request_date = expense.get('request_date')
            if isinstance(request_date, str):
                request_date = datetime.fromisoformat(request_date.replace('Z', '+00:00'))
            ws.cell(row=current_row, column=2, value=request_date)
            ws.cell(row=current_row, column=2).number_format = 'DD.MM.YYYY'

            # Category info
            category = expense.get('category', {})
            ws.cell(row=current_row, column=3, value=category.get('name', ''))
            ws.cell(row=current_row, column=4, value=category.get('type', ''))

            # Contractor
            contractor = expense.get('contractor', {})
            ws.cell(row=current_row, column=5, value=contractor.get('name', '') if contractor else '')

            # Organization
            organization = expense.get('organization', {})
            ws.cell(row=current_row, column=6, value=organization.get('name', ''))

            # Amount
            amount = float(expense.get('amount', 0))
            ws.cell(row=current_row, column=7, value=amount)
            ws.cell(row=current_row, column=7).number_format = '#,##0.00 ₽'

            # Status
            ws.cell(row=current_row, column=8, value=expense.get('status', ''))

            # Payment date
            payment_date = expense.get('payment_date')
            if payment_date:
                if isinstance(payment_date, str):
                    payment_date = datetime.fromisoformat(payment_date.replace('Z', '+00:00'))
                ws.cell(row=current_row, column=9, value=payment_date)
                ws.cell(row=current_row, column=9).number_format = 'DD.MM.YYYY'

            # Comment
            ws.cell(row=current_row, column=10, value=expense.get('comment', ''))

            # Apply borders
            for col in range(1, 11):
                ws.cell(row=current_row, column=col).border = ExcelExporter.BORDER

            # Apply color coding based on category type
            if category.get('type') == 'OPEX':
                for col in range(1, 11):
                    ws.cell(row=current_row, column=col).fill = ExcelExporter.OPEX_FILL
            elif category.get('type') == 'CAPEX':
                for col in range(1, 11):
                    ws.cell(row=current_row, column=col).fill = ExcelExporter.CAPEX_FILL

            current_row += 1

        # Total row
        current_row += 1
        ws.cell(row=current_row, column=6, value='ИТОГО:')
        ws.cell(row=current_row, column=6).font = ExcelExporter.TOTAL_FONT
        ws.cell(row=current_row, column=6).alignment = Alignment(horizontal='right')

        total_col = get_column_letter(7)
        ws.cell(row=current_row, column=7, value=f"=SUM({total_col}4:{total_col}{current_row-2})")
        ws.cell(row=current_row, column=7).number_format = '#,##0.00 ₽'
        ws.cell(row=current_row, column=7).font = ExcelExporter.TOTAL_FONT

        for col in range(6, 8):
            ws.cell(row=current_row, column=col).fill = ExcelExporter.TOTAL_FILL
            ws.cell(row=current_row, column=col).border = ExcelExporter.BORDER

        # Adjust column widths
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 14
        ws.column_dimensions['J'].width = 30

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    @staticmethod
    def export_budget_overview(year: int, month: int, overview_data: Dict[str, Any]) -> BytesIO:
        """
        Export budget overview (Plan vs Actual) to Excel

        Args:
            year: Year
            month: Month
            overview_data: Budget overview data

        Returns:
            BytesIO: Excel file in memory
        """
        wb = Workbook()
        ws = wb.active
        ws.title = f"{year}-{month:02d}"

        # Header
        month_names = ['', 'Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
                      'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']
        ws['A1'] = f"Обзор бюджета: {month_names[month]} {year}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:G1')

        # Column headers
        headers = ['Статья расходов', 'Тип', 'План', 'Факт', 'Остаток', '% исполнения', 'Статус']

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num, value=header)
            cell.font = ExcelExporter.HEADER_FONT
            cell.fill = ExcelExporter.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = ExcelExporter.BORDER

        # Data rows
        current_row = 4
        categories = overview_data.get('categories', [])

        for cat in categories:
            ws.cell(row=current_row, column=1, value=cat['category_name'])
            ws.cell(row=current_row, column=2, value=cat['category_type'])
            ws.cell(row=current_row, column=3, value=cat['planned'])
            ws.cell(row=current_row, column=3).number_format = '#,##0.00'
            ws.cell(row=current_row, column=4, value=cat['actual'])
            ws.cell(row=current_row, column=4).number_format = '#,##0.00'
            ws.cell(row=current_row, column=5, value=cat['remaining'])
            ws.cell(row=current_row, column=5).number_format = '#,##0.00'
            ws.cell(row=current_row, column=6, value=cat['execution_percent'] / 100)
            ws.cell(row=current_row, column=6).number_format = '0.00%'

            # Status
            status = 'Перерасход' if cat['is_overspent'] else 'OK'
            ws.cell(row=current_row, column=7, value=status)

            # Color code remaining column
            remaining = cat['remaining']
            planned = cat['planned']
            if remaining < 0:
                fill_color = "FFC7CE"  # Red
            elif planned > 0 and (remaining / planned) < 0.2:
                fill_color = "FFEB9C"  # Yellow
            else:
                fill_color = "C6EFCE"  # Green

            ws.cell(row=current_row, column=5).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

            # Apply borders
            for col in range(1, 8):
                ws.cell(row=current_row, column=col).border = ExcelExporter.BORDER

            current_row += 1

        # Totals section
        current_row += 1
        totals = overview_data.get('totals', {})
        opex_totals = overview_data.get('opex_totals', {})
        capex_totals = overview_data.get('capex_totals', {})

        for label, data in [('ИТОГО OPEX', opex_totals), ('ИТОГО CAPEX', capex_totals), ('ВСЕГО', totals)]:
            ws.cell(row=current_row, column=1, value=label)
            ws.cell(row=current_row, column=1).font = ExcelExporter.TOTAL_FONT
            ws.cell(row=current_row, column=3, value=data.get('planned', 0))
            ws.cell(row=current_row, column=3).number_format = '#,##0.00'
            ws.cell(row=current_row, column=4, value=data.get('actual', 0))
            ws.cell(row=current_row, column=4).number_format = '#,##0.00'
            ws.cell(row=current_row, column=5, value=data.get('remaining', 0))
            ws.cell(row=current_row, column=5).number_format = '#,##0.00'
            ws.cell(row=current_row, column=6, value=data.get('execution_percent', 0) / 100)
            ws.cell(row=current_row, column=6).number_format = '0.00%'

            for col in range(1, 8):
                ws.cell(row=current_row, column=col).fill = ExcelExporter.TOTAL_FILL
                ws.cell(row=current_row, column=col).border = ExcelExporter.BORDER
                ws.cell(row=current_row, column=col).font = ExcelExporter.TOTAL_FONT

            current_row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    @staticmethod
    def export_forecast_calendar(year: int, month: int, forecasts: List[Dict[str, Any]]) -> BytesIO:
        """
        Export forecast data as calendar (dates in columns)

        Args:
            year: Year
            month: Month (1-12)
            forecasts: List of forecast items with amounts grouped by day

        Returns:
            BytesIO: Excel file in memory
        """
        from calendar import monthrange, day_name
        from datetime import date

        wb = Workbook()
        ws = wb.active

        # Month names in Russian
        month_names = [
            '', 'ЯНВАРЬ', 'ФЕВРАЛЬ', 'МАРТ', 'АПРЕЛЬ', 'МАЙ', 'ИЮНЬ',
            'ИЮЛЬ', 'АВГУСТ', 'СЕНТЯБРЬ', 'ОКТЯБРЬ', 'НОЯБРЬ', 'ДЕКАБРЬ'
        ]

        # Day names in Russian (Monday=0)
        day_names_ru = ['пн', 'вт', 'ср', 'чт', 'пт', 'сб', 'вс']

        ws.title = f"{month_names[month]} {year}"

        # Row 2: "суммы с НДС"
        ws['A2'] = 'суммы с НДС'
        ws['A2'].font = Font(italic=True, size=10)

        # Row 3: Title
        ws['A3'] = f"{month_names[month]}_ПЛАН ПЛАТЕЖЕЙ"
        ws['A3'].font = Font(bold=True, size=12)
        ws.merge_cells('A3:F3')

        # Get all days in month
        _, num_days = monthrange(year, month)
        dates = []
        for day in range(1, num_days + 1):
            d = date(year, month, day)
            dates.append(d)

        # Row 4: Day names (пн, вт, ср...)
        for i, d in enumerate(dates):
            col = 7 + i  # Start from column G
            day_name_short = day_names_ru[d.weekday()]
            cell = ws.cell(row=4, column=col, value=day_name_short)
            cell.font = Font(size=9)
            cell.alignment = Alignment(horizontal='center')

            # Highlight weekends
            if d.weekday() >= 5:  # Saturday or Sunday
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        # Row 5: Headers
        headers = ['N п/п', 'Статья ДДС', 'ЮЛ', 'Контрагент', 'Договор', 'Комментарии']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_num, value=header)
            cell.font = ExcelExporter.HEADER_FONT
            cell.fill = ExcelExporter.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = ExcelExporter.BORDER

        # Date headers
        for i, d in enumerate(dates):
            col = 7 + i
            cell = ws.cell(row=5, column=col, value=d.day)
            cell.font = ExcelExporter.HEADER_FONT
            cell.fill = ExcelExporter.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = ExcelExporter.BORDER

            # Highlight weekends
            if d.weekday() >= 5:
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        # Group forecasts by unique combination
        grouped_forecasts = {}
        for forecast in forecasts:
            category = forecast.get('category') or {}
            contractor = forecast.get('contractor') or {}
            organization = forecast.get('organization') or {}

            category_name = category.get('name', 'Без категории')
            contractor_name = contractor.get('name', '')
            org_name = organization.get('name', '')

            key = (category_name, org_name, contractor_name)

            if key not in grouped_forecasts:
                grouped_forecasts[key] = {
                    'category': category_name,
                    'organization': org_name,
                    'contractor': contractor_name,
                    'amounts_by_date': {}
                }

            # Add amount to the specific date
            forecast_date = forecast.get('date')
            if forecast_date:
                if isinstance(forecast_date, str):
                    forecast_date = datetime.fromisoformat(forecast_date.replace('Z', '+00:00')).date()
                amount = float(forecast.get('amount', 0))

                if forecast_date not in grouped_forecasts[key]['amounts_by_date']:
                    grouped_forecasts[key]['amounts_by_date'][forecast_date] = 0
                grouped_forecasts[key]['amounts_by_date'][forecast_date] += amount

        # Data rows
        current_row = 6
        row_number = 1

        for key, data in sorted(grouped_forecasts.items()):
            # Row number
            ws.cell(row=current_row, column=1, value=row_number)
            ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')

            # Category
            ws.cell(row=current_row, column=2, value=data['category'])

            # Organization
            ws.cell(row=current_row, column=3, value=data['organization'])

            # Contractor
            ws.cell(row=current_row, column=4, value=data['contractor'])

            # Contract (empty for now)
            ws.cell(row=current_row, column=5, value='')

            # Comment (empty for now)
            ws.cell(row=current_row, column=6, value='')

            # Fill amounts by date
            row_total = 0
            for i, d in enumerate(dates):
                col = 7 + i
                amount = data['amounts_by_date'].get(d, 0)

                if amount > 0:
                    ws.cell(row=current_row, column=col, value=amount)
                    ws.cell(row=current_row, column=col).number_format = '#,##0.00'
                    row_total += amount
                else:
                    ws.cell(row=current_row, column=col, value=None)

                ws.cell(row=current_row, column=col).border = ExcelExporter.BORDER
                ws.cell(row=current_row, column=col).alignment = Alignment(horizontal='right')

                # Highlight weekends
                if d.weekday() >= 5:
                    ws.cell(row=current_row, column=col).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

            # Apply borders to all cells in row
            for col in range(1, 7):
                ws.cell(row=current_row, column=col).border = ExcelExporter.BORDER

            current_row += 1
            row_number += 1

        # Totals row
        ws.cell(row=current_row, column=1, value='')
        ws.cell(row=current_row, column=2, value='ИТОГО')
        ws.cell(row=current_row, column=2).font = ExcelExporter.TOTAL_FONT

        for i in range(len(dates)):
            col = 7 + i
            col_letter = get_column_letter(col)
            formula = f"=SUM({col_letter}6:{col_letter}{current_row-1})"
            ws.cell(row=current_row, column=col, value=formula)
            ws.cell(row=current_row, column=col).number_format = '#,##0.00'
            ws.cell(row=current_row, column=col).font = ExcelExporter.TOTAL_FONT
            ws.cell(row=current_row, column=col).fill = ExcelExporter.TOTAL_FILL
            ws.cell(row=current_row, column=col).border = ExcelExporter.BORDER

        for col in range(1, 7):
            ws.cell(row=current_row, column=col).fill = ExcelExporter.TOTAL_FILL
            ws.cell(row=current_row, column=col).border = ExcelExporter.BORDER

        # Adjust column widths
        ws.column_dimensions['A'].width = 8   # N п/п
        ws.column_dimensions['B'].width = 30  # Статья ДДС
        ws.column_dimensions['C'].width = 20  # ЮЛ
        ws.column_dimensions['D'].width = 25  # Контрагент
        ws.column_dimensions['E'].width = 15  # Договор
        ws.column_dimensions['F'].width = 20  # Комментарии

        # Date columns
        for i in range(len(dates)):
            col_letter = get_column_letter(7 + i)
            ws.column_dimensions[col_letter].width = 10

        # Freeze panes at column G (after Комментарии)
        ws.freeze_panes = 'G6'

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    @staticmethod
    def export_forecast_from_template(
        year: int,
        month: int,
        forecasts: List[Dict[str, Any]],
        department_name: str = "Шикунов",
        template_path: Optional[Path] = None
    ) -> BytesIO:
        """
        Export forecast data using existing template file

        Args:
            year: Year
            month: Month (1-12)
            forecasts: List of forecast items
            department_name: Name of the sheet/department (default: "Шикунов" for IT)
            template_path: Path to template file (optional, uses default if not provided)

        Returns:
            BytesIO: Excel file in memory
        """
        # Определяем путь к шаблону
        if template_path is None:
            # Используем путь относительно корня проекта
            template_path = Path(__file__).parent.parent.parent.parent / "xls" / "Планирование_10.2025-3.xlsx"

        if not template_path.exists():
            raise FileNotFoundError(f"Шаблон не найден: {template_path}")

        # Загружаем шаблон
        wb = load_workbook(template_path)

        # Проверяем наличие нужного листа
        if department_name not in wb.sheetnames:
            raise ValueError(f"Лист '{department_name}' не найден в шаблоне. Доступные листы: {wb.sheetnames}")

        ws = wb[department_name]

        # Определяем количество дней в месяце
        _, num_days = monthrange(year, month)

        # Группируем прогнозы по уникальной комбинации
        grouped_forecasts = {}
        for forecast in forecasts:
            category = forecast.get('category') or {}
            contractor = forecast.get('contractor') or {}
            organization = forecast.get('organization') or {}

            category_name = category.get('name', 'Без категории')
            contractor_name = contractor.get('name', '')
            org_name = organization.get('name', '')
            comment = forecast.get('comment', '')

            key = (category_name, org_name, contractor_name)

            if key not in grouped_forecasts:
                grouped_forecasts[key] = {
                    'category': category_name,
                    'organization': org_name,
                    'contractor': contractor_name,
                    'comment': comment,
                    'amounts_by_day': {}
                }

            # Добавляем сумму к конкретному дню
            forecast_date = forecast.get('date')
            if forecast_date:
                if isinstance(forecast_date, str):
                    forecast_date = datetime.fromisoformat(forecast_date.replace('Z', '+00:00')).date()
                amount = float(forecast.get('amount', 0))
                day = forecast_date.day

                if day not in grouped_forecasts[key]['amounts_by_day']:
                    grouped_forecasts[key]['amounts_by_day'][day] = 0
                grouped_forecasts[key]['amounts_by_day'][day] += amount

        # Находим начало данных (строка 5 - это заголовки, данные начинаются со строки 6 в шаблоне)
        # Но мы будем искать первую пустую строку после строки 4 или использовать строку 5+
        data_start_row = 5

        # Очищаем старые данные (если есть) начиная со строки data_start_row
        # Сначала найдем последнюю заполненную строку
        last_row = ws.max_row
        for row in range(data_start_row, last_row + 1):
            # Очищаем данные в колонках B-F (Статья, ЮЛ, Контрагент, Договор, Комментарии)
            for col in range(2, 7):  # B=2, C=3, D=4, E=5, F=6
                ws.cell(row=row, column=col).value = None
            # Очищаем суммы по дням (колонки H и далее, H=8)
            for col in range(8, 8 + num_days):
                ws.cell(row=row, column=col).value = None

        # Заполняем данные прогноза
        current_row = data_start_row
        row_number = 1

        for key, data in sorted(grouped_forecasts.items()):
            # Номер п/п (колонка B)
            ws.cell(row=current_row, column=2, value=row_number)

            # Статья ДДС (колонка C)
            ws.cell(row=current_row, column=3, value=data['category'])

            # ЮЛ (колонка D)
            ws.cell(row=current_row, column=4, value=data['organization'])

            # Контрагент (колонка E)
            ws.cell(row=current_row, column=5, value=data['contractor'])

            # Договор (колонка F) - оставляем пустым
            ws.cell(row=current_row, column=6, value='')

            # Комментарии (колонка G)
            ws.cell(row=current_row, column=7, value=data['comment'])

            # Заполняем суммы по дням (начиная с колонки H = 8)
            for day in range(1, num_days + 1):
                col = 7 + day  # H=8, I=9, J=10, ...
                amount = data['amounts_by_day'].get(day, 0)

                if amount > 0:
                    ws.cell(row=current_row, column=col, value=amount)
                    ws.cell(row=current_row, column=col).number_format = '#,##0.00'
                else:
                    ws.cell(row=current_row, column=col, value=None)

            current_row += 1
            row_number += 1

        # Сохраняем в BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output
