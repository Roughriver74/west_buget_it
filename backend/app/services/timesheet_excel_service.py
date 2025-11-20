"""
Timesheet Excel export/import service
"""

from io import BytesIO
from typing import List, Dict, Any, Optional
from datetime import datetime, date
from calendar import monthrange
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.db.models import Employee, WorkTimesheet, DailyWorkRecord


class TimesheetExcelService:
    """Service for timesheet Excel operations"""

    # Styling
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

    WEEKEND_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    HOLIDAY_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    WORKED_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    TOTAL_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    TOTAL_FONT = Font(bold=True, size=11)

    BORDER = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    @staticmethod
    def export_timesheet_grid(
        year: int,
        month: int,
        employees_data: List[Dict[str, Any]],
        department_name: str = "Отдел"
    ) -> BytesIO:
        """
        Export timesheet grid to Excel

        Args:
            year: Year
            month: Month (1-12)
            employees_data: List of employee timesheet data
            department_name: Department name for sheet title

        Returns:
            BytesIO: Excel file in memory
        """
        wb = Workbook()
        ws = wb.active

        # Russian month names
        month_names = [
            '', 'Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
            'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь'
        ]

        ws.title = f"{month_names[month]} {year}"

        # Title
        ws['A1'] = f"Табель учета рабочего времени: {department_name}, {month_names[month]} {year}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells(f'A1:{get_column_letter(5 + monthrange(year, month)[1])}1')

        # Get number of days in month
        _, num_days = monthrange(year, month)

        # Column headers
        headers = ['№', 'Сотрудник', 'Должность', 'Табельный номер']

        # Add day columns
        for day in range(1, num_days + 1):
            headers.append(str(day))

        # Add totals columns
        headers.extend(['Всего дней', 'Всего часов', 'Статус'])

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num, value=header)
            cell.font = TimesheetExcelService.HEADER_FONT
            cell.fill = TimesheetExcelService.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = TimesheetExcelService.BORDER

        # Helper to check if day is weekend
        def is_weekend(year: int, month: int, day: int) -> bool:
            day_of_week = date(year, month, day).weekday()
            return day_of_week in (5, 6)  # Saturday or Sunday

        # Data rows
        current_row = 4
        for idx, emp_data in enumerate(employees_data, 1):
            # Employee info
            ws.cell(row=current_row, column=1, value=idx)
            ws.cell(row=current_row, column=2, value=emp_data.get('employee_full_name', ''))
            ws.cell(row=current_row, column=3, value=emp_data.get('employee_position', ''))
            ws.cell(row=current_row, column=4, value=emp_data.get('employee_number', ''))

            # Daily hours
            days = emp_data.get('days', [])
            for day in range(1, num_days + 1):
                col = 4 + day
                day_data = days[day - 1] if day - 1 < len(days) else None
                hours = day_data.get('hours_worked', 0) if day_data else 0

                cell = ws.cell(row=current_row, column=col, value=hours if hours > 0 else None)
                cell.alignment = Alignment(horizontal='center')
                cell.border = TimesheetExcelService.BORDER

                # Color code weekends
                if is_weekend(year, month, day):
                    cell.fill = TimesheetExcelService.WEEKEND_FILL
                elif hours > 0:
                    cell.fill = TimesheetExcelService.WORKED_FILL

            # Totals
            total_days = emp_data.get('total_days_worked', 0)
            total_hours = emp_data.get('total_hours_worked', 0)
            status = emp_data.get('timesheet_status', 'Не создан')

            ws.cell(row=current_row, column=num_days + 5, value=total_days)
            ws.cell(row=current_row, column=num_days + 5).alignment = Alignment(horizontal='center')
            ws.cell(row=current_row, column=num_days + 5).font = Font(bold=True)
            ws.cell(row=current_row, column=num_days + 5).border = TimesheetExcelService.BORDER

            ws.cell(row=current_row, column=num_days + 6, value=total_hours)
            ws.cell(row=current_row, column=num_days + 6).alignment = Alignment(horizontal='center')
            ws.cell(row=current_row, column=num_days + 6).font = Font(bold=True)
            ws.cell(row=current_row, column=num_days + 6).border = TimesheetExcelService.BORDER

            # Status with Russian labels
            status_labels = {
                'DRAFT': 'Черновик',
                'APPROVED': 'Утвержден',
                'PAID': 'Оплачен'
            }
            ws.cell(row=current_row, column=num_days + 7, value=status_labels.get(status, status))
            ws.cell(row=current_row, column=num_days + 7).alignment = Alignment(horizontal='center')
            ws.cell(row=current_row, column=num_days + 7).border = TimesheetExcelService.BORDER

            # Apply borders to employee info columns
            for col in range(1, 5):
                ws.cell(row=current_row, column=col).border = TimesheetExcelService.BORDER

            current_row += 1

        # Summary row
        current_row += 1
        ws.cell(row=current_row, column=2, value='ИТОГО')
        ws.cell(row=current_row, column=2).font = TimesheetExcelService.TOTAL_FONT

        # Calculate totals for each day
        for day in range(1, num_days + 1):
            col = 4 + day
            col_letter = get_column_letter(col)
            formula = f"=SUM({col_letter}4:{col_letter}{current_row-2})"
            cell = ws.cell(row=current_row, column=col, value=formula)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = TimesheetExcelService.TOTAL_FILL
            cell.font = TimesheetExcelService.TOTAL_FONT
            cell.border = TimesheetExcelService.BORDER

            # Color weekends
            if is_weekend(year, month, day):
                cell.fill = TimesheetExcelService.WEEKEND_FILL

        # Total days and hours
        col_days = num_days + 5
        col_hours = num_days + 6

        ws.cell(row=current_row, column=col_days, value=f"=SUM({get_column_letter(col_days)}4:{get_column_letter(col_days)}{current_row-2})")
        ws.cell(row=current_row, column=col_days).fill = TimesheetExcelService.TOTAL_FILL
        ws.cell(row=current_row, column=col_days).font = TimesheetExcelService.TOTAL_FONT
        ws.cell(row=current_row, column=col_days).border = TimesheetExcelService.BORDER

        ws.cell(row=current_row, column=col_hours, value=f"=SUM({get_column_letter(col_hours)}4:{get_column_letter(col_hours)}{current_row-2})")
        ws.cell(row=current_row, column=col_hours).fill = TimesheetExcelService.TOTAL_FILL
        ws.cell(row=current_row, column=col_hours).font = TimesheetExcelService.TOTAL_FONT
        ws.cell(row=current_row, column=col_hours).border = TimesheetExcelService.BORDER

        # Apply total fill to info columns
        for col in range(1, 5):
            ws.cell(row=current_row, column=col).fill = TimesheetExcelService.TOTAL_FILL
            ws.cell(row=current_row, column=col).border = TimesheetExcelService.BORDER

        # Adjust column widths
        ws.column_dimensions['A'].width = 5   # №
        ws.column_dimensions['B'].width = 30  # Сотрудник
        ws.column_dimensions['C'].width = 25  # Должность
        ws.column_dimensions['D'].width = 15  # Табельный номер

        # Day columns
        for day in range(1, num_days + 1):
            col_letter = get_column_letter(4 + day)
            ws.column_dimensions[col_letter].width = 5

        # Totals columns
        ws.column_dimensions[get_column_letter(num_days + 5)].width = 12
        ws.column_dimensions[get_column_letter(num_days + 6)].width = 12
        ws.column_dimensions[get_column_letter(num_days + 7)].width = 12

        # Freeze panes (freeze first 4 columns)
        ws.freeze_panes = 'E4'

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    @staticmethod
    def generate_timesheet_template(
        year: int,
        month: int,
        employees: List[Employee],
        language: str = 'ru'
    ) -> BytesIO:
        """
        Generate empty timesheet template for manual filling

        Args:
            year: Year
            month: Month (1-12)
            employees: List of employees
            language: Template language ('ru' or 'en')

        Returns:
            BytesIO: Excel template file
        """
        wb = Workbook()
        ws = wb.active

        # Month names
        if language == 'ru':
            month_names = [
                '', 'Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
                'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь'
            ]
            ws.title = f"Шаблон {month_names[month]}"
            headers = ['№', 'Табельный номер', 'ФИО', 'Должность']
        else:
            month_names = [
                '', 'January', 'February', 'March', 'April', 'May', 'June',
                'July', 'August', 'September', 'October', 'November', 'December'
            ]
            ws.title = f"Template {month_names[month]}"
            headers = ['#', 'Employee Number', 'Full Name', 'Position']

        # Title
        title = f"{'Шаблон табеля' if language == 'ru' else 'Timesheet Template'}: {month_names[month]} {year}"
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=14)

        # Get number of days
        _, num_days = monthrange(year, month)
        ws.merge_cells(f'A1:{get_column_letter(4 + num_days)}1')

        # Add day columns to headers
        for day in range(1, num_days + 1):
            headers.append(str(day))

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num, value=header)
            cell.font = TimesheetExcelService.HEADER_FONT
            cell.fill = TimesheetExcelService.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = TimesheetExcelService.BORDER

        # Helper for weekend check
        def is_weekend(year: int, month: int, day: int) -> bool:
            return date(year, month, day).weekday() in (5, 6)

        # Color weekend columns
        for day in range(1, num_days + 1):
            if is_weekend(year, month, day):
                col = 4 + day
                ws.cell(row=3, column=col).fill = TimesheetExcelService.WEEKEND_FILL

        # Add employee rows
        current_row = 4
        for idx, emp in enumerate(employees, 1):
            ws.cell(row=current_row, column=1, value=idx)
            ws.cell(row=current_row, column=2, value=emp.employee_number)
            ws.cell(row=current_row, column=3, value=emp.full_name)
            ws.cell(row=current_row, column=4, value=emp.position)

            # Apply borders and weekend coloring
            for col in range(1, 5):
                ws.cell(row=current_row, column=col).border = TimesheetExcelService.BORDER

            # Empty cells for days with weekend highlighting
            for day in range(1, num_days + 1):
                col = 4 + day
                cell = ws.cell(row=current_row, column=col, value=None)
                cell.border = TimesheetExcelService.BORDER
                cell.alignment = Alignment(horizontal='center')

                if is_weekend(year, month, day):
                    cell.fill = TimesheetExcelService.WEEKEND_FILL

            current_row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 25

        for day in range(1, num_days + 1):
            ws.column_dimensions[get_column_letter(4 + day)].width = 5

        # Add instructions
        current_row += 2
        if language == 'ru':
            ws[f'A{current_row}'] = "ИНСТРУКЦИЯ:"
            ws[f'A{current_row}'].font = Font(bold=True, size=12)
            current_row += 1
            ws[f'A{current_row}'] = "1. Заполните количество отработанных часов для каждого сотрудника по дням месяца"
            current_row += 1
            ws[f'A{current_row}'] = "2. Серым цветом отмечены выходные дни"
            current_row += 1
            ws[f'A{current_row}'] = "3. Введите 0 или оставьте пустым для неотработанных дней"
            current_row += 1
            ws[f'A{current_row}'] = "4. После заполнения загрузите файл через систему импорта"
        else:
            ws[f'A{current_row}'] = "INSTRUCTIONS:"
            ws[f'A{current_row}'].font = Font(bold=True, size=12)
            current_row += 1
            ws[f'A{current_row}'] = "1. Fill in worked hours for each employee by day of month"
            current_row += 1
            ws[f'A{current_row}'] = "2. Gray cells indicate weekends"
            current_row += 1
            ws[f'A{current_row}'] = "3. Enter 0 or leave empty for non-working days"
            current_row += 1
            ws[f'A{current_row}'] = "4. After filling, upload the file through the import system"

        # Freeze panes
        ws.freeze_panes = 'E4'

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output
