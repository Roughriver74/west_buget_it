"""
Excel Template Generator

Generates Excel templates with proper formatting and examples
based on import configurations.
"""

from typing import Optional, List, Dict, Any
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.services.import_config_manager import ImportConfig, get_import_config_manager


class TemplateGenerator:
    """Generate Excel templates for import"""

    def __init__(self):
        self.config_manager = get_import_config_manager()

        # Style definitions
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, size=11)
        self.required_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
        self.optional_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        self.example_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        self.border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

    def generate_template(
        self,
        entity_type: str,
        language: str = "ru",
        include_examples: bool = True,
        include_instructions: bool = True
    ) -> Optional[BytesIO]:
        """
        Generate Excel template for entity

        Args:
            entity_type: Entity type (e.g., "employees")
            language: Language code ("ru" or "en")
            include_examples: Include example rows
            include_instructions: Include instructions sheet

        Returns:
            BytesIO with Excel file or None if entity not found
        """
        config = self.config_manager.get_config(entity_type)
        if not config:
            return None

        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create main data sheet
        ws_data = wb.create_sheet(title="Данные" if language == "ru" else "Data")
        self._generate_data_sheet(ws_data, config, language, include_examples)

        # Create instructions sheet
        if include_instructions:
            ws_instructions = wb.create_sheet(title="Инструкция" if language == "ru" else "Instructions")
            self._generate_instructions_sheet(ws_instructions, config, language)

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    def _generate_data_sheet(
        self,
        ws,
        config: ImportConfig,
        language: str,
        include_examples: bool
    ):
        """Generate main data sheet with headers and examples"""
        # Start with header row immediately (no empty rows before)
        # This is required for auto-detection of columns during import
        header_row = 1
        for col_idx, field in enumerate(config.fields, start=1):
            col_letter = get_column_letter(col_idx)
            cell = ws[f"{col_letter}{header_row}"]

            # Set column name
            display_name = field["display_name"].get(language, field["name"])
            required_marker = " *" if field.get("required", False) else ""
            cell.value = f"{display_name}{required_marker}"

            # Style header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border

            # Set column width
            ws.column_dimensions[col_letter].width = self._get_column_width(field)

        ws.row_dimensions[header_row].height = 40

        # Type hint row
        type_row = header_row + 1
        for col_idx, field in enumerate(config.fields, start=1):
            col_letter = get_column_letter(col_idx)
            cell = ws[f"{col_letter}{type_row}"]

            # Get type info
            type_info = self._get_type_hint(field, language)
            cell.value = type_info

            # Style based on required/optional
            if field.get("required", False):
                cell.fill = self.required_fill
            else:
                cell.fill = self.optional_fill

            cell.font = Font(size=9, italic=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border

        ws.row_dimensions[type_row].height = 30

        # Example rows
        if include_examples:
            examples = self._get_example_rows(config, language)
            start_row = type_row + 1

            for row_idx, example in enumerate(examples, start=start_row):
                for col_idx, field in enumerate(config.fields, start=1):
                    col_letter = get_column_letter(col_idx)
                    cell = ws[f"{col_letter}{row_idx}"]

                    cell.value = example.get(field["name"], "")
                    cell.fill = self.example_fill
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.border = self.border

                ws.row_dimensions[row_idx].height = 20

        # Freeze panes (header and type rows)
        ws.freeze_panes = f"A{type_row + 1}"

    def _generate_instructions_sheet(
        self,
        ws,
        config: ImportConfig,
        language: str
    ):
        """Generate instructions sheet"""
        instructions = {
            "ru": {
                "title": "Инструкция по импорту",
                "required": "Обязательные поля",
                "optional": "Необязательные поля",
                "formats": "Форматы данных",
                "tips": "Советы",
                "required_text": "Поля отмеченные * обязательны для заполнения",
                "optional_text": "Остальные поля необязательны",
                "date_format": "Даты: ДД.ММ.ГГГГ (например: 31.12.2024)",
                "bool_format": "Логические значения: Да/Нет, 1/0, True/False",
                "number_format": "Числа: используйте точку как разделитель (например: 1234.56)",
                "tip1": "Не удаляйте строку с названиями колонок",
                "tip2": "Удалите примеры перед импортом своих данных",
                "tip3": "Сохраняйте файл в формате .xlsx или .xls",
                "tip4": "Проверьте данные перед импортом",
            },
            "en": {
                "title": "Import Instructions",
                "required": "Required Fields",
                "optional": "Optional Fields",
                "formats": "Data Formats",
                "tips": "Tips",
                "required_text": "Fields marked with * are required",
                "optional_text": "Other fields are optional",
                "date_format": "Dates: DD.MM.YYYY (example: 31.12.2024)",
                "bool_format": "Boolean: Yes/No, 1/0, True/False",
                "number_format": "Numbers: use dot as separator (example: 1234.56)",
                "tip1": "Do not delete the header row",
                "tip2": "Remove examples before importing your data",
                "tip3": "Save file as .xlsx or .xls",
                "tip4": "Validate data before import",
            }
        }

        texts = instructions.get(language, instructions["ru"])

        # Title
        ws['A1'] = texts["title"]
        ws['A1'].font = Font(size=16, bold=True, color="4472C4")
        ws.row_dimensions[1].height = 25

        row = 3

        # Required fields
        ws[f'A{row}'] = texts["required"]
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1

        required_fields = config.get_required_fields()
        for field_name in required_fields:
            field = config.get_field(field_name)
            if field:
                display_name = field["display_name"].get(language, field_name)
                ws[f'A{row}'] = f"• {display_name}"
                row += 1

        row += 1

        # Optional fields
        ws[f'A{row}'] = texts["optional"]
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        ws[f'A{row}'] = texts["optional_text"]
        row += 2

        # Formats
        ws[f'A{row}'] = texts["formats"]
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1

        for field in config.fields:
            field_type = field.get("type")
            display_name = field["display_name"].get(language, field["name"])

            if field_type == "date":
                ws[f'A{row}'] = f"• {display_name}: {texts['date_format']}"
                row += 1
            elif field_type == "boolean":
                ws[f'A{row}'] = f"• {display_name}: {texts['bool_format']}"
                row += 1
            elif field_type in ["decimal", "integer"]:
                ws[f'A{row}'] = f"• {display_name}: {texts['number_format']}"
                row += 1
            elif field_type == "enum":
                enum_values = field.get("enum", [])
                ws[f'A{row}'] = f"• {display_name}: {', '.join(enum_values)}"
                row += 1

        row += 1

        # Tips
        ws[f'A{row}'] = texts["tips"]
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1

        tips = ["tip1", "tip2", "tip3", "tip4"]
        for tip in tips:
            ws[f'A{row}'] = f"✓ {texts[tip]}"
            row += 1

        # Set column width
        ws.column_dimensions['A'].width = 80

    def _get_column_width(self, field: Dict[str, Any]) -> int:
        """Calculate optimal column width"""
        field_type = field.get("type")

        if field_type == "boolean":
            return 12
        elif field_type in ["integer", "decimal"]:
            return 15
        elif field_type == "date":
            return 15
        elif field_type == "enum":
            return 20
        else:
            max_length = field.get("max_length", 50)
            return min(max(max_length // 2, 20), 50)

    def _get_type_hint(self, field: Dict[str, Any], language: str) -> str:
        """Get type hint text for field"""
        field_type = field.get("type")

        type_hints = {
            "ru": {
                "string": "Текст",
                "integer": "Целое число",
                "decimal": "Число",
                "boolean": "Да/Нет",
                "date": "ДД.ММ.ГГГГ",
                "enum": "Выбор",
            },
            "en": {
                "string": "Text",
                "integer": "Integer",
                "decimal": "Number",
                "boolean": "Yes/No",
                "date": "DD.MM.YYYY",
                "enum": "Choice",
            }
        }

        hints = type_hints.get(language, type_hints["ru"])
        base_hint = hints.get(field_type, "")

        # Add enum values
        if field_type == "enum":
            enum_values = field.get("enum", [])
            if enum_values:
                base_hint += f": {', '.join(enum_values[:3])}"
                if len(enum_values) > 3:
                    base_hint += "..."

        # Add min/max for numbers
        if field_type in ["integer", "decimal"]:
            min_val = field.get("min")
            max_val = field.get("max")
            if min_val is not None or max_val is not None:
                range_text = []
                if min_val is not None:
                    range_text.append(f"≥{min_val}")
                if max_val is not None:
                    range_text.append(f"≤{max_val}")
                base_hint += f" ({', '.join(range_text)})"

        return base_hint

    def _get_example_rows(self, config: ImportConfig, language: str) -> List[Dict[str, Any]]:
        """Generate example rows"""
        examples_ru = {
            "budget_categories": [
                {"name": "Оборудование", "type": "CAPEX", "description": "Серверы, компьютеры", "is_active": "Да"},
                {"name": "Интернет", "type": "OPEX", "description": "Услуги связи", "is_active": "Да"},
            ],
            "contractors": [
                {"name": "ООО Компания А", "short_name": "Компания А", "inn": "1234567890", "contact_info": "+7-999-999-99-99", "is_active": "Да"},
                {"name": "ИП Иванов И.И.", "short_name": "Иванов", "inn": "123456789012", "contact_info": "ivanov@mail.ru", "is_active": "Да"},
            ],
            "employees": [
                {"full_name": "Иванов Иван Иванович", "position": "Системный администратор", "base_salary": 80000, "employee_number": "E001", "birth_date": "15.06.1985", "monthly_bonus_base": 8000, "quarterly_bonus_base": 24000, "annual_bonus_base": 0, "hire_date": "01.01.2023"},
                {"full_name": "Петрова Мария Сергеевна", "position": "Разработчик", "base_salary": 120000, "employee_number": "E002", "birth_date": "22.09.1990", "monthly_bonus_base": 12000, "quarterly_bonus_base": 0, "annual_bonus_base": 50000, "hire_date": "15.03.2023"},
            ],
            "payroll_plans": [
                {"year": 2025, "month": 1, "employee_name": "Иванов Иван Иванович", "base_salary": 80000, "bonus": 8000, "other_payments": 0, "notes": ""},
                {"year": 2025, "month": 1, "employee_name": "Петрова Мария Сергеевна", "base_salary": 120000, "bonus": 12000, "other_payments": 5000, "notes": "Квартальная премия"},
            ],
            "expenses": [
                {"number": "EXP-2025-001", "category_name": "Интернет", "contractor_name": "ООО Провайдер", "organization_name": "ООО Наша компания", "amount": 15000, "request_date": "01.01.2025", "payment_date": "", "status": "PENDING", "purpose": "Оплата интернета за январь", "comment": "", "requester": "Иванов И.И."},
                {"number": "EXP-2025-002", "category_name": "Оборудование", "contractor_name": "ООО Техника", "organization_name": "ООО Наша компания", "amount": 250000, "request_date": "05.01.2025", "payment_date": "10.01.2025", "status": "PAID", "purpose": "Покупка сервера", "comment": "Оплачено", "requester": "Петров П.П."},
            ],
            "budget_plans": [
                {"year": 2025, "month": 1, "category_name": "Интернет", "planned_amount": 15000, "capex_planned": 0, "opex_planned": 15000, "status": "DRAFT"},
                {"year": 2025, "month": 1, "category_name": "Оборудование", "planned_amount": 250000, "capex_planned": 250000, "opex_planned": 0, "status": "APPROVED"},
                {"year": 2025, "month": 2, "category_name": "Интернет", "planned_amount": 15000, "capex_planned": 0, "opex_planned": 15000, "status": "DRAFT"},
            ],
            "budget_plan_details": [
                {"version_id": 1, "category_name": "Интернет", "type": "OPEX", "month": 1, "planned_amount": 15000, "subcategory": "", "calculation_method": "AVERAGE", "business_driver": "Количество сотрудников", "justification": "Средний расход за прошлый год", "based_on_year": 2024, "growth_rate": 5.0},
                {"version_id": 1, "category_name": "Оборудование", "type": "CAPEX", "month": 1, "planned_amount": 250000, "subcategory": "Серверы", "calculation_method": "MANUAL", "business_driver": "Новый проект", "justification": "Покупка сервера для нового сервиса", "based_on_year": None, "growth_rate": None},
            ],
            "revenue_plan_details": [
                {"version_id": 1, "revenue_stream_name": "Подписки", "revenue_category_name": "SaaS", "month_01": 100000, "month_02": 105000, "month_03": 110000, "month_04": 115000, "month_05": 120000, "month_06": 125000, "month_07": 130000, "month_08": 135000, "month_09": 140000, "month_10": 145000, "month_11": 150000, "month_12": 155000},
                {"version_id": 1, "revenue_stream_name": "Консалтинг", "revenue_category_name": "Услуги", "month_01": 50000, "month_02": 50000, "month_03": 60000, "month_04": 60000, "month_05": 70000, "month_06": 70000, "month_07": 80000, "month_08": 80000, "month_09": 90000, "month_10": 90000, "month_11": 100000, "month_12": 100000},
            ],
        }

        examples_en = {
            "budget_categories": [
                {"name": "Equipment", "type": "CAPEX", "description": "Servers, computers", "is_active": "Yes"},
                {"name": "Internet", "type": "OPEX", "description": "Communication services", "is_active": "Yes"},
            ],
            "contractors": [
                {"name": "Company A LLC", "short_name": "Company A", "inn": "1234567890", "contact_info": "+7-999-999-99-99", "is_active": "Yes"},
                {"name": "IP Ivanov I.I.", "short_name": "Ivanov", "inn": "123456789012", "contact_info": "ivanov@mail.ru", "is_active": "Yes"},
            ],
            "employees": [
                {"full_name": "John Smith", "position": "System Administrator", "base_salary": 80000, "employee_number": "E001", "birth_date": "15.06.1985", "monthly_bonus_base": 8000, "quarterly_bonus_base": 24000, "annual_bonus_base": 0, "hire_date": "01.01.2023"},
                {"full_name": "Mary Johnson", "position": "Developer", "base_salary": 120000, "employee_number": "E002", "birth_date": "22.09.1990", "monthly_bonus_base": 12000, "quarterly_bonus_base": 0, "annual_bonus_base": 50000, "hire_date": "15.03.2023"},
            ],
            "payroll_plans": [
                {"year": 2025, "month": 1, "employee_name": "John Smith", "base_salary": 80000, "bonus": 8000, "other_payments": 0, "notes": ""},
                {"year": 2025, "month": 1, "employee_name": "Mary Johnson", "base_salary": 120000, "bonus": 12000, "other_payments": 5000, "notes": "Quarterly bonus"},
            ],
            "expenses": [
                {"number": "EXP-2025-001", "category_name": "Internet", "contractor_name": "Provider LLC", "organization_name": "Our Company LLC", "amount": 15000, "request_date": "01.01.2025", "payment_date": "", "status": "PENDING", "purpose": "Internet payment for January", "comment": "", "requester": "J. Smith"},
                {"number": "EXP-2025-002", "category_name": "Equipment", "contractor_name": "Tech LLC", "organization_name": "Our Company LLC", "amount": 250000, "request_date": "05.01.2025", "payment_date": "10.01.2025", "status": "PAID", "purpose": "Server purchase", "comment": "Paid", "requester": "P. Johnson"},
            ],
            "budget_plans": [
                {"year": 2025, "month": 1, "category_name": "Internet", "planned_amount": 15000, "capex_planned": 0, "opex_planned": 15000, "status": "DRAFT"},
                {"year": 2025, "month": 1, "category_name": "Equipment", "planned_amount": 250000, "capex_planned": 250000, "opex_planned": 0, "status": "APPROVED"},
                {"year": 2025, "month": 2, "category_name": "Internet", "planned_amount": 15000, "capex_planned": 0, "opex_planned": 15000, "status": "DRAFT"},
            ],
            "budget_plan_details": [
                {"version_id": 1, "category_name": "Internet", "type": "OPEX", "month": 1, "planned_amount": 15000, "subcategory": "", "calculation_method": "AVERAGE", "business_driver": "Employee count", "justification": "Average expense from last year", "based_on_year": 2024, "growth_rate": 5.0},
                {"version_id": 1, "category_name": "Equipment", "type": "CAPEX", "month": 1, "planned_amount": 250000, "subcategory": "Servers", "calculation_method": "MANUAL", "business_driver": "New project", "justification": "Server purchase for new service", "based_on_year": None, "growth_rate": None},
            ],
            "revenue_plan_details": [
                {"version_id": 1, "revenue_stream_name": "Subscriptions", "revenue_category_name": "SaaS", "month_01": 100000, "month_02": 105000, "month_03": 110000, "month_04": 115000, "month_05": 120000, "month_06": 125000, "month_07": 130000, "month_08": 135000, "month_09": 140000, "month_10": 145000, "month_11": 150000, "month_12": 155000},
                {"version_id": 1, "revenue_stream_name": "Consulting", "revenue_category_name": "Services", "month_01": 50000, "month_02": 50000, "month_03": 60000, "month_04": 60000, "month_05": 70000, "month_06": 70000, "month_07": 80000, "month_08": 80000, "month_09": 90000, "month_10": 90000, "month_11": 100000, "month_12": 100000},
            ],
        }

        examples = examples_ru if language == "ru" else examples_en
        return examples.get(config.entity, [])


# Global instance
_template_generator: Optional[TemplateGenerator] = None


def get_template_generator() -> TemplateGenerator:
    """Get global template generator instance"""
    global _template_generator
    if _template_generator is None:
        _template_generator = TemplateGenerator()
    return _template_generator
